#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
入出力解析 - レポート出力 (I/O Report)

入出力解析結果をExcelワークブック（2シート構成）またはCSVとして出力する。
  シート1: 入出力一覧（ユニットごとの入出力ファイル + source_tag色分け）
  シート2: サマリ（実行パラメタ・件数統計・色凡例）
"""

import csv
import datetime

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def inout_write_excel(path, records, headers, gui_vars=None):
    """入出力解析結果をExcelに出力する（2シート構成: データ + サマリ）"""
    if not OPENPYXL_AVAILABLE: raise ImportError("openpyxl missing")

    FONT_NAME = "ＭＳ ゴシック"
    base_font = Font(name=FONT_NAME, size=10)
    hdr_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal='center', vertical='center')
    wrap_top = Alignment(wrap_text=True, vertical='top')
    bottom_thin = Border(bottom=Side(style='thin', color='FFBFBFBF'))
    # source_tag別の背景色
    fill_fail = PatternFill(start_color="FFFFD2D2", end_color="FFFFD2D2", fill_type="solid")    # 赤: 解析失敗
    fill_exception = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid") # 薄い青: 例外JSON
    fill_no_io = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")   # 薄いグレー: IO定義無
    # サマリ用
    section_font = Font(name=FONT_NAME, size=11, bold=True)
    section_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")

    wb = openpyxl.Workbook()

    # =================================================================
    # シート1: データ
    # =================================================================
    ws = wb.active
    ws.title = "入出力一覧"
    # ヘッダー
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{max(len(records) + 1, 2)}"

    # データ行
    for record in records:
        in_files = "\n".join(record.get('inputs', []))
        out_files = "\n".join(record.get('outputs', []))
        tag = record.get('source_tag', '不明')
        ws.append([
            record.get('unit_full', ''),
            record.get('unit', ''),
            record.get('resource', ''),
            in_files, out_files, tag
        ])
        row_idx = ws.max_row
        # source_tagに応じた色分け
        row_fill = None
        if "解析失敗" in tag:
            row_fill = fill_fail
        elif "例外JSON" in tag or "手動ルール" in tag or "外部ファイル" in tag or "tbl_expand" in tag.lower() or "hulft_tbl" in tag.lower() or "TBL" in tag:
            row_fill = fill_exception
        elif "IO定義無" in tag:
            row_fill = fill_no_io
        for cell in ws[row_idx]:
            cell.font = base_font
            cell.alignment = wrap_top
            cell.border = bottom_thin
            if row_fill:
                cell.fill = row_fill

    # 列幅
    col_widths = {'A': 50, 'B': 18, 'C': 22, 'D': 50, 'E': 50, 'F': 30}
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    # =================================================================
    # シート2: サマリ
    # =================================================================
    ws_sum = wb.create_sheet("サマリ")
    ws_sum.column_dimensions['A'].width = 28
    ws_sum.column_dimensions['B'].width = 70

    def _add_section(label):
        ws_sum.append([label, ""])
        row_idx = ws_sum.max_row
        for cell in ws_sum[row_idx]:
            cell.font = section_font
            cell.fill = section_fill
        ws_sum.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)

    def _add_row(key, val):
        ws_sum.append([key, val])
        for cell in ws_sum[ws_sum.max_row]:
            cell.font = base_font
            cell.alignment = wrap_top
            cell.border = bottom_thin

    # 実行パラメタ
    _add_section("実行パラメタ")
    _add_row("解析日時", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    try:
        _add_row("銀行", gui_vars['v_inout_bank'].get() if gui_vars else "")
    except: _add_row("銀行", "")
    try:
        _add_row("AJSパス", gui_vars['v_inout_ajs'].get() if gui_vars else "")
    except: _add_row("AJSパス", "")

    ws_sum.append(["", ""])

    # 件数サマリ
    total = len(records)
    cnt_fail = sum(1 for r in records if "解析失敗" in r.get('source_tag', ''))
    cnt_exception = sum(1 for r in records if any(k in r.get('source_tag', '') for k in ("例外JSON", "手動ルール", "外部ファイル", "TBL")))
    cnt_no_io = sum(1 for r in records if "IO定義無" in r.get('source_tag', ''))
    cnt_no_resource = sum(1 for r in records if "リソース指定なし" in r.get('source_tag', ''))
    cnt_shell_ok = total - cnt_fail - cnt_exception - cnt_no_io - cnt_no_resource

    _add_section("件数サマリ")
    _add_row("総ユニット数", total)
    _add_row("  シェル解析成功", cnt_shell_ok)
    _add_row("  例外JSON適用", cnt_exception)
    _add_row("  IO定義無", cnt_no_io)
    _add_row("  リソース指定なし", cnt_no_resource)
    _add_row("  解析失敗", cnt_fail)

    ws_sum.append(["", ""])

    # 色凡例
    _add_section("色凡例")
    _add_row("赤", "解析失敗（変数未解決等）")
    _add_row("青", "例外JSON / 手動ルール適用")
    _add_row("グレー", "IO定義無（入出力なしのシェル）")
    _add_row("白", "シェル解析で自動取得")

    wb.save(path)


def inout_write_csv(path, records, headers):
    with open(path, "w", encoding="utf-8-sig", newline="") as cf:
        w = csv.DictWriter(cf, fieldnames=headers)
        w.writeheader()
        for record in records:
            w.writerow({
                headers[0]: record.get('unit_full', ''),
                headers[1]: record.get('unit', ''),
                headers[2]: record.get('resource', ''),
                headers[3]: " | ".join(record.get('inputs', [])),
                headers[4]: " | ".join(record.get('outputs', [])),
                headers[5]: record.get('source_tag', '不明'),
            })
