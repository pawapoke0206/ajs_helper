#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
依存関係解析 - Excelレポート出力 (Dependency Report)

依存関係解析結果をExcelワークブック（4シート構成）として出力する。
  シート1: 入力ファイル調査（外部入力の分類・色分け）
  シート2: 探索ログ（BFS/DFSのトレース記録）
  シート3: 必要ユニット一覧（BFS/充足/ブリッジの区分付き）
  シート4: 結果サマリ（実行パラメタ・統計・項目説明）
"""

import os
import datetime

try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ajs_constants import LOG_DIR
from ajs_utils import make_logger, pre_normalize
from ajs_dep_tracer import find_parent_jobnet

LOG_FILE_RUN = LOG_DIR / "depend_run.log"
_log = make_logger(LOG_FILE_RUN)


def write_dependency_report(
    out_path,
    true_externals,          # {ファイル: {consumer_unit_full, ...}}
    disconnected_externals,  # 同上
    producer_map,            # {ファイル: {producer_unit_full, ...}}
    record_by_full,          # {unit_full: record}
    trace_data,              # [(layer, unit_name, unit_full, input_file, info_str)]
    needed_units_full,       # set
    target_unit_fulls,       # set
    supplemented_units,      # set
    bridged_units,           # set
    not_found,               # list
    ambiguous,               # list
    gui_vars,                # GUI変数（実行パラメタ表示用）
    G=None,                  # NetworkXグラフ
    base_dir_to_remove="",
    parent_jobnets=None,     # 親ジョブネット一覧（クロスジョブネット用）
    choice_cache=None,       # ユーザー選択の履歴（クロスジョブネット用）
    db_producer_map=None,    # {テーブル名 → W元unit_fullの集合}
):
    """依存関係解析結果をExcelレポートとして出力する（4シート構成）"""
    if not OPENPYXL_AVAILABLE:
        _log("[Warning] openpyxl not available, skipping Excel report")
        return

    if parent_jobnets is None:
        parent_jobnets = []
    if db_producer_map is None:
        db_producer_map = {}
    if choice_cache is None:
        choice_cache = {}

    wb = openpyxl.Workbook()
    FONT_NAME = "MS Gothic"

    # --- 共通スタイル定義 ---
    base_font = Font(name=FONT_NAME, size=10)
    hdr_font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal='center', vertical='center')
    section_font = Font(name=FONT_NAME, size=11, bold=True)
    section_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    wrap_top = Alignment(wrap_text=True, vertical='top')
    # 入力ファイル調査の色分け:
    #   外部ファイル（想定出力元なし）→ 最重要、目立つ色（赤系）
    #   外部ファイル（想定出力元あり）→ 次に重要（青系）
    #   TBL等 → あまり見せたくない（暗いグレー）
    fill_ext_no_producer = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")  # 薄い赤/オレンジ
    fill_ext_has_producer = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")  # 薄い青
    fill_tbl = PatternFill(start_color="FFE0E0E0", end_color="FFE0E0E0", fill_type="solid")  # 暗めのグレー
    fill_tbl_font = Font(name=FONT_NAME, size=10, color="FF808080")  # TBL行のフォントも薄く
    # 区切り行
    layer_sep_font = Font(name=FONT_NAME, size=10, bold=True, color="FF305496")
    layer_sep_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    pj_sep_font = Font(name=FONT_NAME, size=11, bold=True, color="FF1F4E79")
    pj_sep_fill = PatternFill(start_color="FFB4C6E7", end_color="FFB4C6E7", fill_type="solid")
    bottom_thin = Border(bottom=Side(style='thin', color='FFBFBFBF'))
    # 区分別の背景色（units用）
    fill_target = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    fill_supplement = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
    fill_bridge = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")

    def _write_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    def _set_col_widths(ws, widths):
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    def _unit_name(unit_full):
        rec = record_by_full.get(unit_full)
        return rec['unit'] if rec else os.path.basename(unit_full)

    def _add_pj_separator(ws, label, col_count):
        """親ジョブネットの区切り行を挿入する"""
        ws.append([label] + [""] * (col_count - 1))
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.font = pj_sep_font
            cell.fill = pj_sep_fill
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=col_count)

    # --- ファイルを親ジョブネットごとに振り分けるヘルパー ---
    def _classify_file(fpath):
        """ファイルの分類を返す: 'tbl', 'master', 'normal'"""
        if '/TBL/' in fpath:
            return 'tbl'
        if fpath in disconnected_externals:
            return 'master'
        return 'normal'

    def _find_file_pj(fpath):
        """ファイルの想定出力元が属する親ジョブネットを返す"""
        producers = producer_map.get(fpath, set())
        producers_in_result = producers & needed_units_full
        for p in producers_in_result:
            pj = find_parent_jobnet(p, parent_jobnets) if parent_jobnets else None
            if pj:
                return pj
        # consumerの親ジョブネットで分類
        consumers = true_externals.get(fpath, set()) | disconnected_externals.get(fpath, set())
        for c in consumers:
            pj = find_parent_jobnet(c, parent_jobnets) if parent_jobnets else None
            if pj:
                return pj
        return None

    # =====================================================================
    # シート1: 入力ファイル調査
    # 親ジョブネットごとにグループ化。TBL系は最下部にまとめる。
    # =====================================================================
    ws_miss = wb.active
    ws_miss.title = "入力ファイル調査"
    _write_header(ws_miss, [
        "ファイルパス", "分類", "使用元ユニット", "使用元（フルパス）",
        "想定出力元", "マスタ判定", "備考"])
    _set_col_widths(ws_miss, {
        'A': 48, 'B': 14, 'C': 24, 'D': 48, 'E': 48, 'F': 10, 'G': 16})

    def _norm_path(unit_full):
        """unit_fullをbase_dir_to_removeで正規化して表示用パスにする"""
        return pre_normalize(unit_full, base_dir_to_remove) if unit_full else ""

    def _add_file_row(ws, fpath, category, consumers, suggested="",
                      is_master=False, remark="", row_fill=None, use_tbl_font=False):
        consumer_names = ", ".join(sorted(_unit_name(c) for c in consumers))
        consumer_fulls = "\n".join(sorted(_norm_path(c) for c in consumers))
        suggested_disp = _norm_path(suggested)
        master_mark = "\u25CB" if is_master else ""
        ws.append([fpath, category, consumer_names, consumer_fulls,
                   suggested_disp, master_mark, remark])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.font = fill_tbl_font if use_tbl_font else base_font
            cell.alignment = wrap_top
            cell.border = bottom_thin
            if row_fill:
                cell.fill = row_fill

    # 全外部入力ファイルを分類
    all_ext_files = set(true_externals.keys()) | set(disconnected_externals.keys())
    tbl_files = []      # TBL系（最下部にまとめる）
    pj_files = {}       # {親JNパス: [entry, ...]}
    no_pj_files = []    # 親JN不明

    def _get_suggested_producer(fpath, consumers=None):
        """ファイルの想定出力元をフルパスで返す。
        DFSで解決されたproducer（choice_cacheに記録済み）のみを返す。
        充足チェックで自動追加された同一親JN内のproducerは返さない。"""
        if consumers and choice_cache:
            for c in sorted(consumers):
                chosen = choice_cache.get((fpath, c))
                if chosen:
                    return chosen
        return ""

    for fpath in sorted(all_ext_files):
        all_consumers = true_externals.get(fpath, set()) | disconnected_externals.get(fpath, set())
        is_disc = fpath in disconnected_externals  # マスタ系（producerは存在するが未接続だった）
        # TBLDIR配下（/SAM01/TBL/）のみTBL扱い。skats系等は除外。
        is_tbl = '/SAM01/TBL/' in fpath or '/SAM02/TBL/' in fpath
        is_db = fpath.startswith("[DB]")  # DB依存（テーブル名）

        if is_tbl:
            tbl_files.append((fpath, all_consumers))
            continue

        # consumersを親JNごとに振り分け
        consumers_by_pj = {}
        consumers_no_pj = set()
        for c in all_consumers:
            c_pj = find_parent_jobnet(c, parent_jobnets) if parent_jobnets else None
            if c_pj:
                consumers_by_pj.setdefault(c_pj, set()).add(c)
            else:
                consumers_no_pj.add(c)

        def _make_entry(fpath, consumers_set, consumer_pj=None):
            suggested = _get_suggested_producer(fpath, consumers_set)
            # A列表示用: DB系は[DB]プレフィックスを除去
            display_path = fpath[4:] if is_db else fpath
            # 色: 想定出力元の有無で決定
            if suggested:
                fill = fill_ext_has_producer
            else:
                fill = fill_ext_no_producer

            if is_db:
                # DB系の分類: 同じ親JN内にW元があれば「内部更新」
                db_table_name = fpath[4:]  # [DB]を除去
                is_internal = False
                if consumer_pj and parent_jobnets:
                    db_producers = db_producer_map.get(db_table_name, set())
                    for p in db_producers:
                        if find_parent_jobnet(p, parent_jobnets) == consumer_pj:
                            is_internal = True
                            break
                elif is_disc:
                    is_internal = True
                if is_internal:
                    category = "外部DB（内部更新）"
                else:
                    category = "外部DB"
                is_master = is_internal
            else:
                # ファイル系の分類（既存ロジック）
                # マスタ判定: この親JN内にproducerが存在するかどうか
                is_master = False
                if consumer_pj and parent_jobnets:
                    producers = producer_map.get(fpath, set())
                    for p in producers:
                        if find_parent_jobnet(p, parent_jobnets) == consumer_pj:
                            is_master = True
                            break
                elif is_disc:
                    is_master = True
                if is_master:
                    category = "外部ファイル (マスタ系)"
                else:
                    category = "外部ファイル"
            return (display_path, consumers_set, suggested, is_master, category, fill)

        for pj, pj_consumers in consumers_by_pj.items():
            entry = _make_entry(fpath, pj_consumers, consumer_pj=pj)
            pj_files.setdefault(pj, []).append(entry)

        if consumers_no_pj:
            entry = _make_entry(fpath, consumers_no_pj, consumer_pj=None)
            no_pj_files.append(entry)

    # 親JNごとに出力
    for pj_path in sorted(pj_files.keys()):
        pj_name = os.path.basename(pj_path)
        _add_pj_separator(ws_miss, pj_name, 7)
        for fpath, consumers, suggested, is_master, category, fill in pj_files[pj_path]:
            _add_file_row(ws_miss, fpath, category, consumers,
                          suggested=suggested, is_master=is_master,
                          row_fill=fill)

    # 親JN不明
    if no_pj_files:
        for fpath, consumers, suggested, is_master, category, fill in no_pj_files:
            _add_file_row(ws_miss, fpath, category, consumers,
                          suggested=suggested, is_master=is_master,
                          row_fill=fill)

    # TBL等（最下部にまとめる、暗い色+薄いフォント）
    if tbl_files:
        _add_pj_separator(ws_miss, "TBL\u7B49", 7)
        for fpath, consumers in tbl_files:
            _add_file_row(ws_miss, fpath, "TBL\u7B49", consumers,
                          row_fill=fill_tbl, use_tbl_font=True)

    # =====================================================================
    # シート2: 探索ログ（DFS+BFS入れ子構造）
    # trace_dataにはクロスジョブネットのセクション情報も含まれている
    # =====================================================================
    ws_trace = wb.create_sheet("探索ログ")
    _write_header(ws_trace, [
        "Layer", "ユニット名", "ユニット（フルパス）", "入力ファイル", "判定結果"])
    _set_col_widths(ws_trace, {'A': 8, 'B': 22, 'C': 48, 'D': 48, 'E': 60})

    prev_layer = None
    prev_unit_name = None
    prev_unit_full = None
    for layer, unit_name, unit_full, input_file, info_str in trace_data:
        # DFSセクション区切り行（layer="SECTION"のマーカー）
        if layer == "SECTION":
            ws_trace.append([info_str, "", "", "", ""])
            sep_row = ws_trace.max_row
            for cell in ws_trace[sep_row]:
                cell.font = pj_sep_font
                cell.fill = pj_sep_fill
            ws_trace.merge_cells(start_row=sep_row, start_column=1,
                                 end_row=sep_row, end_column=5)
            prev_layer = None
            prev_unit_name = None
            prev_unit_full = None
            continue

        # BFSレイヤー区切り行
        if layer != prev_layer:
            ws_trace.append([f"Layer {layer}", "", "", "", ""])
            sep_row = ws_trace.max_row
            for cell in ws_trace[sep_row]:
                cell.font = layer_sep_font
                cell.fill = layer_sep_fill
            ws_trace.merge_cells(start_row=sep_row, start_column=1,
                                 end_row=sep_row, end_column=5)
            prev_layer = layer
            prev_unit_name = None
            prev_unit_full = None

        is_new_unit = (unit_name != prev_unit_name or unit_full != prev_unit_full)
        disp_layer = layer if is_new_unit else ""
        disp_name = unit_name if is_new_unit else ""
        disp_full = unit_full if is_new_unit else ""
        prev_unit_name = unit_name
        prev_unit_full = unit_full

        ws_trace.append([disp_layer, disp_name, disp_full, input_file, info_str])
        for cell in ws_trace[ws_trace.max_row]:
            cell.font = base_font
            cell.alignment = wrap_top
            cell.border = bottom_thin

    # =====================================================================
    # シート3: 必要ユニット一覧（親JN列追加）
    # =====================================================================
    ws_units = wb.create_sheet("必要ユニット一覧")
    _write_header(ws_units, [
        "ユニット名", "フルパス", "親ジョブネット", "区分",
        "入力ファイル", "出力ファイル"])
    _set_col_widths(ws_units, {
        'A': 22, 'B': 48, 'C': 24, 'D': 14, 'E': 48, 'F': 48})

    def _unit_io(unit_full):
        rec = record_by_full.get(unit_full)
        if not rec:
            return "", ""
        inputs = "\n".join(f for f in rec.get('inputs', []) if f)
        outputs = "\n".join(f for f in rec.get('outputs', []) if f)
        return inputs, outputs

    def _add_unit_row(ws, unit_full, category, row_fill=None):
        inputs, outputs = _unit_io(unit_full)
        pj = find_parent_jobnet(unit_full, parent_jobnets) if parent_jobnets else None
        pj_name = os.path.basename(pj) if pj else ""
        ws.append([_unit_name(unit_full), unit_full, pj_name, category,
                   inputs, outputs])
        for cell in ws[ws.max_row]:
            cell.font = base_font
            cell.alignment = wrap_top
            cell.border = bottom_thin
            if row_fill:
                cell.fill = row_fill

    bfs_only = needed_units_full - supplemented_units - bridged_units
    for unit_full in sorted(bfs_only):
        is_target = unit_full in target_unit_fulls
        category = "\u76EE\u6A19" if is_target else "BFS"
        _add_unit_row(ws_units, unit_full, category,
                      row_fill=fill_target if is_target else None)

    for unit_full in sorted(supplemented_units):
        _add_unit_row(ws_units, unit_full, "\u5145\u8DB3\u30C1\u30A7\u30C3\u30AF",
                      row_fill=fill_supplement)

    for unit_full in sorted(bridged_units):
        _add_unit_row(ws_units, unit_full, "DB\u30D6\u30EA\u30C3\u30B8",
                      row_fill=fill_bridge)

    # =====================================================================
    # シート4: 結果サマリ
    # =====================================================================
    ws_sum = wb.create_sheet("結果サマリ")
    _set_col_widths(ws_sum, {'A': 28, 'B': 70})

    def _add_section_header(ws, label):
        ws.append([label, ""])
        row_idx = ws.max_row
        for cell in ws[row_idx]:
            cell.font = section_font
            cell.fill = section_fill
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=2)

    def _add_kv_row(ws, key, val):
        ws.append([key, val])
        for cell in ws[ws.max_row]:
            cell.font = base_font
            cell.alignment = wrap_top
            cell.border = bottom_thin

    # 実行パラメタ
    try:
        ajs_path_val = gui_vars.get('v_dep_ajs', gui_vars.get('v_inout_ajs', None))
        ajs_path_val = ajs_path_val.get() if ajs_path_val else ""
    except Exception:
        ajs_path_val = ""
    try:
        bank_val = gui_vars.get('v_dep_bank', gui_vars.get('v_inout_bank', None))
        bank_val = bank_val.get() if bank_val else ""
    except Exception:
        bank_val = ""
    target_str = "\n".join(sorted(target_unit_fulls))

    _add_section_header(ws_sum, "\u5B9F\u884C\u30D1\u30E9\u30E1\u30BF")
    _add_kv_row(ws_sum, "\u89E3\u6790\u65E5\u6642",
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    _add_kv_row(ws_sum, "AJS\u30D1\u30B9", ajs_path_val)
    _add_kv_row(ws_sum, "\u9280\u884C", bank_val)
    _add_kv_row(ws_sum, "\u76EE\u6A19\u30E6\u30CB\u30C3\u30C8", target_str)
    ws_sum.append(["", ""])

    # 結果サマリ
    bfs_count = len(needed_units_full) - len(supplemented_units) - len(bridged_units)
    _add_section_header(ws_sum, "\u7D50\u679C\u30B5\u30DE\u30EA")
    _add_kv_row(ws_sum, "\u76EE\u6A19\u30E6\u30CB\u30C3\u30C8\u6570",
                len(target_unit_fulls))
    _add_kv_row(ws_sum, "\u5FC5\u8981\u30B8\u30E7\u30D6\u6570\uFF08\u5408\u8A08\uFF09",
                len(needed_units_full))
    _add_kv_row(ws_sum, "  \u3046\u3061BFS\u63A2\u7D22", bfs_count)
    if supplemented_units:
        _add_kv_row(ws_sum, "  \u3046\u3061\u5145\u8DB3\u30C1\u30A7\u30C3\u30AF\u8FFD\u52A0",
                    len(supplemented_units))
    if bridged_units:
        _add_kv_row(ws_sum, "  \u3046\u3061DB\u30D6\u30EA\u30C3\u30B8\u8FFD\u52A0",
                    len(bridged_units))
    all_ext = set(true_externals.keys()) | set(disconnected_externals.keys())
    _add_kv_row(ws_sum, "\u5916\u90E8\u5165\u529B\u30D5\u30A1\u30A4\u30EB\u6570\uFF08\u5408\u8A08\uFF09",
                len(all_ext))
    _add_kv_row(ws_sum, "  \u5916\u90E8\u30D5\u30A1\u30A4\u30EB\uFF08TBL\u7B49\uFF09",
                len(true_externals))
    _add_kv_row(ws_sum, "  \u30DE\u30B9\u30BF\u7CFB", len(disconnected_externals))
    ws_sum.append(["", ""])

    # クロスジョブネット情報
    if parent_jobnets and len(parent_jobnets) > 1:
        _add_section_header(ws_sum,
                            "\u30AF\u30ED\u30B9\u30B8\u30E7\u30D6\u30CD\u30C3\u30C8\u60C5\u5831")
        _add_kv_row(ws_sum,
                    "\u89AA\u30B8\u30E7\u30D6\u30CD\u30C3\u30C8\u6570",
                    len(parent_jobnets))
        for i, pj in enumerate(sorted(parent_jobnets), 1):
            _add_kv_row(ws_sum, f"  [{i}]", os.path.basename(pj))
        if choice_cache:
            ws_sum.append(["", ""])
            _add_kv_row(ws_sum,
                        "\u30E6\u30FC\u30B6\u30FC\u9078\u629E\u5C65\u6B74",
                        f"{len(choice_cache)}\u4EF6")
            for (fpath, consumer), chosen in sorted(choice_cache.items(),
                                                    key=lambda x: x[0]):
                consumer_name = os.path.basename(consumer) if consumer else ""
                _add_kv_row(ws_sum,
                            f"  {os.path.basename(fpath)} ({consumer_name})",
                            os.path.basename(chosen))
            ws_sum.append(["", ""])

    # 項目説明
    _add_section_header(ws_sum, "\u9805\u76EE\u8AAC\u660E")
    _add_kv_row(ws_sum, "BFS\u63A2\u7D22",
                "\u76EE\u6A19\u30E6\u30CB\u30C3\u30C8\u304B\u3089\u5165\u529B\u30D5\u30A1\u30A4\u30EB\u306E\u4F5C\u6210\u5143\u3092\u518D\u5E30\u7684\u306B\u305F\u3069\u3063\u3066\u898B\u3064\u3051\u305F\u30B8\u30E7\u30D6\u6570")
    _add_kv_row(ws_sum, "\u5145\u8DB3\u30C1\u30A7\u30C3\u30AF\u8FFD\u52A0",
                "BFS\u63A2\u7D22\u3060\u3051\u3067\u306F\u6F0F\u308C\u305F\u30B8\u30E7\u30D6\u3092\u88DC\u5B8C\u3057\u305F\u3082\u306E")
    _add_kv_row(ws_sum, "DB\u30D6\u30EA\u30C3\u30B8\u8FFD\u52A0",
                "DB\u7D4C\u7531\u306E\u4F9D\u5B58\u95A2\u4FC2\u7B49\u306B\u3088\u308A\u8FFD\u52A0\u3055\u308C\u305F\u30B8\u30E7\u30D6\u6570")
    _add_kv_row(ws_sum, "\u30DE\u30B9\u30BF\u7CFB",
                "\u4F5C\u6210\u5143\u30B8\u30E7\u30D6\u306F\u5B58\u5728\u3059\u308B\u304C\u7D50\u679C\u306B\u542B\u307E\u308C\u306A\u304B\u3063\u305F\u30D5\u30A1\u30A4\u30EB\u3002\u65E5\u6B21\u66F4\u65B0\u7CFB\u306E\u30DE\u30B9\u30BF\u7B49\u304C\u8A72\u5F53")
    ws_sum.append(["", ""])

    # 注意事項
    _add_section_header(ws_sum, "\u6CE8\u610F\u4E8B\u9805")
    if not_found:
        _add_kv_row(ws_sum, "\u898B\u3064\u304B\u3089\u306A\u304B\u3063\u305F\u30E6\u30CB\u30C3\u30C8",
                    ", ".join(not_found))
    if ambiguous:
        _add_kv_row(ws_sum, "\u540C\u540D\u3067\u7279\u5B9A\u4E0D\u80FD",
                    ", ".join(t for t, _ in ambiguous))
    if not not_found and not ambiguous:
        _add_kv_row(ws_sum, "", "\u6CE8\u610F\u4E8B\u9805\u306A\u3057")

    # --- 保存 ---
    wb.save(out_path)
    _log(f"[Info] Dependency report saved: {out_path}")
