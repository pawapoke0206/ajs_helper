#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AJS Helper Tool - 入出力解析ロジック (In/Out Logic)
v5.5 (2025-11-23) - 変数名リファクタリング, SVNAME削除, ログ機能強化, 出力構成変更(tmp/固定名)
"""

import re
import os
import sys
import time
import glob
import codecs
import shlex
import csv
import json
import pathlib
import itertools
import copy
import traceback
import datetime
from fnmatch import fnmatch

# openpyxl は Excel 出力に必要
try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 定数ファイルをインポート
from ajs_constants import LOG_DIR, IO_EXCEPTION_FILE, CONFIG_FILE, DIR_NAME_INOUT

# --- グローバル変数: 解析結果のキャッシュ ---
_ANALYSIS_CACHE = {}
_LAST_CACHE_KEY = None

# ログファイルパス
LOG_FILE_RUN = LOG_DIR / "tab3_inout_run.log"
LOG_FILE_DETAIL = LOG_DIR / "tab3_inout_details.json"

def _log(msg):
    """実行ログへの書き込みヘルパー"""
    try:
        timestamp = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        with open(LOG_FILE_RUN, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {msg}\n")
    except: pass

# -----------------------------------------------------------------------------
# ★ グローバル正規表現定義
# -----------------------------------------------------------------------------
RES_PAT = re.compile(r"^(FILE([IO])(\d{2})|CBL_SYS([01])(\d{2})|([IO])(\d{2})FILE|IN_FILE|OUT_FILE)=(.+)$", re.VERBOSE)
ALL_VAR_PAT = re.compile(r"\$(\{([^}]+)\}|([a-zA-Z_][a-zA-Z0-9_]*)|(\d+))")
CASE_VAR_PAT = re.compile(r'case\s*["\']?\$(\{([^}]+)\}|([a-zA-Z_][a-zA-Z0-9_]*))["\']?')
VAR_ASSIGN_PAT = re.compile(r'^\s*(?:export\s+)?([^=\s]+)\s*=\s*(' r'[\'"](.*?)[\'"]|' r'(?:[^ \t;#]+)' r')')
RM_PAT = re.compile(r'^\s*rm\s+(?:-f\s+)?(.*)')

# -----------------------------------------------------------------------------
# ★ ヘルパー関数 & クラス
# -----------------------------------------------------------------------------
def _create_replacer(var_dict):
    def replacer(match):
        key = match.group(2) or match.group(3) or match.group(4)
        return var_dict.get(key, match.group(0)) if key else match.group(0)
    return replacer

class ComenvParser:
    def __init__(self, comenv_path, initial_vars, log_data):
        self.comenv_path = comenv_path
        self.initial_vars = initial_vars
        self.log_data = log_data 
        self.lines = self._read_comenv()
        self.case_patterns = {} 
        self.master_var_dict = {} 
        self.case_vars_order = [] 

    def _read_comenv(self):
        if not self.comenv_path: return []
        try:
            with codecs.open(self.comenv_path, "r", "cp932", errors="ignore") as f:
                return f.readlines()
        except Exception as e:
            _log(f"[Warning] comenv read error: {e}")
            return []

    def _resolve_value(self, value, current_vars):
        if "$" not in value: return value
        replacer_func = _create_replacer(current_vars)
        for _ in range(10): 
            new_value = ALL_VAR_PAT.sub(replacer_func, value)
            if new_value == value: return new_value 
            value = new_value
        return value

    def _evaluate_if_statement(self, line, current_vars):
        match = re.search(r'if\s*\[\s*"\$\{([^}]+)\}"\s*=\s*"([^"]+)"\s*\]', line)
        if match:
            return current_vars.get(match.group(1), "") == match.group(2)
        return True

    def parse_all_patterns(self):
        if not self.lines: return
        
        # case文の変数を抽出
        current_case_var = None
        for line in self.lines:
            line_stripped = line.strip()
            if line_stripped.startswith("case"):
                match = CASE_VAR_PAT.search(line_stripped)
                if match:
                    current_case_var = match.group(2) or match.group(3)
                    if current_case_var:
                        if current_case_var not in self.case_vars_order:
                            self.case_vars_order.append(current_case_var)
                        self.case_patterns.setdefault(current_case_var, set())
            elif current_case_var:
                match = re.match(r'^\s*([^)\s]+)\s*\)', line_stripped)
                if match:
                    pattern = match.group(1).strip()
                    if pattern != "*": self.case_patterns[current_case_var].add(pattern)
            elif line_stripped.startswith("esac"):
                current_case_var = None
        
        # パターン組み合わせ生成
        pattern_lists = []
        for var_name in self.case_vars_order:
            patterns = sorted(list(self.case_patterns.get(var_name, set())))
            patterns.append("*") 
            pattern_lists.append(patterns)
        
        if not pattern_lists:
            pattern_lists = [["*"]]
            self.case_vars_order = ["_default_"] 
            
        self.log_data["comenv_case_patterns"] = self.case_patterns
        all_combinations = list(itertools.product(*pattern_lists))
        
        for combination in all_combinations:
            current_pattern_map = dict(zip(self.case_vars_order, combination)) 
            current_vars = self.initial_vars.copy()
            active_block = True      
            case_active = False    
            current_case_var = None    
            case_match_found = False   
            
            for line in self.lines:
                line_stripped = line.strip()
                if not line_stripped or line_stripped.startswith("#"): continue
                if line_stripped.startswith("if"):
                    active_block = self._evaluate_if_statement(line_stripped, current_vars)
                    continue
                elif line_stripped.startswith("else"):
                    active_block = not active_block
                    continue
                elif line_stripped.startswith("fi"):
                    active_block = True
                    continue

                if line_stripped.startswith("case"):
                    match = CASE_VAR_PAT.search(line_stripped)
                    if match:
                        current_case_var = match.group(2) or match.group(3)
                        case_match_found = False
                        case_active = True 
                    continue
                elif line_stripped.startswith("esac"):
                    current_case_var = None
                    case_match_found = False
                    case_active = False
                    active_block = True 
                    continue
                
                is_pattern_line = False
                if case_active:
                    match = re.match(r'^\s*([^)\s]+)\s*\)', line_stripped)
                    if match:
                        is_pattern_line = True
                        pattern = match.group(1).strip()
                        target_pattern = current_pattern_map.get(current_case_var, "*")
                        if not case_match_found and (pattern == target_pattern or pattern == "*"):
                            active_block = True
                            case_match_found = True 
                        else:
                            active_block = False
                    elif line_stripped.startswith(";;"):
                        active_block = False 
                        continue
                
                if active_block:
                    line_to_parse = line.replace("export ", "").strip()
                    if is_pattern_line: line_to_parse = line_to_parse.split(")", 1)[-1].strip()
                    if line_to_parse.endswith(";;"): line_to_parse = line_to_parse[:-2].strip()
                    line_to_parse = line_to_parse.split('#', 1)[0].strip()
                    var_match = VAR_ASSIGN_PAT.match(line_to_parse)
                    if var_match:
                        var_name = var_match.group(1)
                        value = var_match.group(3) if var_match.group(3) is not None else var_match.group(2)
                        value = value.strip("'\"")
                        current_vars[var_name] = self._resolve_value(value, current_vars)
            self.master_var_dict[combination] = current_vars

    def get_var_dict_for_env(self, ajs_env_str):
        if not self.master_var_dict: return self.initial_vars 
        env_vars = dict(item.split('=', 1) for item in ajs_env_str.split(';') if '=' in item)
        key_tuple = []
        for var_name in self.case_vars_order:
            val = env_vars.get(var_name, "*") 
            if var_name in self.case_patterns and val not in self.case_patterns[var_name]: val = "*"
            key_tuple.append(val)
        return self.master_var_dict.get(tuple(key_tuple), self.initial_vars)

class ShellParser:
    def __init__(self, shell_path):
        self.shell_path = shell_path
        self.procedures = []
        self._parse_shell_to_procedures()

    def _parse_shell_to_procedures(self):
        try:
            with codecs.open(self.shell_path, "r", "cp932", errors="ignore") as f:
                for line in f:
                    line = line.strip().split('#', 1)[0].strip()
                    if not line: continue
                    rm_match = RM_PAT.match(line)
                    if rm_match:
                        self.procedures.append(("RM", rm_match.group(1).strip()))
                        continue 
                    a_match = VAR_ASSIGN_PAT.match(line)
                    if a_match:
                        val = a_match.group(3) if a_match.group(3) is not None else a_match.group(2)
                        self.procedures.append(("ASSIGN", a_match.group(1), val.strip("'\"")))
                    b_match = RES_PAT.match(line)
                    if b_match:
                        var = b_match.group(0).split('=',1)[0]
                        val = b_match.group(8)
                        io_groups = (b_match.group(1), b_match.group(2), b_match.group(4), b_match.group(6))
                        self.procedures.append(("IO_ASSIGN", var, val.strip("'\""), io_groups))
        except Exception as e:
            _log(f"[Warning] Shell parse error ({self.shell_path}): {e}")

    def get_procedures(self): return self.procedures

class ShellExecutor:
    def __init__(self, procedures, comenv_dict, ajs_record):
        self.procedures = procedures
        self.comenv_dict = comenv_dict
        self.ajs_record = ajs_record
        self.shell_context = {} 
        self._init_context()

    def _init_context(self):
        self.shell_context = copy.deepcopy(self.comenv_dict)
        try:
            env_vars = dict(item.split('=', 1) for item in self.ajs_record['env'].split(';') if '=' in item)
            self.shell_context.update(env_vars)
        except Exception: pass 
        try:
            params = self.ajs_record['param'].split()
            for i, param in enumerate(params): self.shell_context[f"{i+1}"] = param 
        except Exception: pass 

    def _resolve_value(self, value_template):
        if not value_template or "$" not in value_template: return value_template
        replacer_func = _create_replacer(self.shell_context)
        return ALL_VAR_PAT.sub(replacer_func, value_template)

    def execute(self):
        inputs, outputs = [], []
        unresolved_io_vars = set() 
        for proc in self.procedures:
            op = proc[0]
            if op == "RM":
                resolved = self._resolve_value(proc[1])
                outputs = [f for f in outputs if f != resolved]
                inputs = [f for f in inputs if f != resolved]
                continue
            name, val_tmpl = proc[1], proc[2]
            resolved_val = self._resolve_value(val_tmpl)
            if op == "ASSIGN":
                self.shell_context[name] = resolved_val
            elif op == "IO_ASSIGN":
                self.shell_context[name] = resolved_val 
                is_unresolved = False
                if ALL_VAR_PAT.findall(resolved_val):
                    for mt in ALL_VAR_PAT.findall(resolved_val):
                        key = mt[1] or mt[2] or mt[3]
                        if key not in self.shell_context: is_unresolved = True
                if '`' in resolved_val or '$(' in resolved_val: is_unresolved = True
                if is_unresolved: unresolved_io_vars.add(val_tmpl) 
                tag, io2, sys01, io3 = proc[3]
                is_input = False
                if "IN_FILE" in tag: is_input = True
                elif "OUT_FILE" in tag: is_input = False
                elif io2: is_input = (io2 == "I")
                elif sys01: is_input = (sys01 == "0")
                elif io3: is_input = (io3 == "I")
                if is_input: inputs.append(resolved_val)
                else: outputs.append(resolved_val)
        return inputs, outputs, list(unresolved_io_vars)

def inout_parse_ini_resource(path: str):
    inputs, outputs = [], []
    try:
        with codecs.open(path, "r", "cp932", errors="ignore") as f:
            for line in f:
                m = RES_PAT.match(line.strip())
                if not m: continue
                val = m.group(8).strip()
                tag, io2, sys01, io3 = m.group(1), m.group(2), m.group(4), m.group(6)
                is_input = False
                if "IN_FILE" in tag: is_input = True
                elif "OUT_FILE" in tag: is_input = False
                elif io2: is_input = (io2 == "I")
                elif sys01: is_input = (sys01 == "0")
                elif io3: is_input = (io3 == "I")
                if is_input: inputs.append(val)
                else: outputs.append(val)
    except Exception: pass 
    return sorted(list(set(inputs))), sorted(list(set(outputs)))

def write_detail_log(log_data):
    """詳細ログ(JSON)出力"""
    try:
        ld = log_data.copy()
        # JSONシリアライズできないSet型などを変換
        if "comenv_case_patterns" in ld:
            ld["comenv_case_patterns"] = {k: list(v) for k, v in ld["comenv_case_patterns"].items()}
        if "comenv_master_dictionary" in ld:
            ld["comenv_master_dictionary"] = {",".join(k): v for k, v in ld["comenv_master_dictionary"].items()}
        
        with open(LOG_FILE_DETAIL, "w", encoding="utf-8") as f:
            json.dump(ld, f, indent=2, ensure_ascii=False)
        _log(f"[Info] Detail log saved: {LOG_FILE_DETAIL}")
    except Exception as e:
        _log(f"[Error] Failed to write detail log: {e}")

def inout_parse_ajsprint_output(local_tmp_path):
    ajs_mapping_list = []
    headers = ["unit_full", "unit", "resource", "type", "env", "param"]
    with codecs.open(local_tmp_path, "r", "cp932", errors="ignore") as f:
        for line in f:
            parts = line.rstrip("\n").split("\t")
            if len(parts) < len(headers): parts.extend([""] * (len(headers) - len(parts)))
            ajs_mapping_list.append(dict(zip(headers, parts)))
    return ajs_mapping_list

def inout_resolve_path_variables(paths, var_dict):
    resolved_paths, unresolved_vars = [], set()
    replacer_func = _create_replacer(var_dict)
    for path in paths:
        if "$" not in path:
            resolved_paths.append(path)
            continue
        resolved_path = ALL_VAR_PAT.sub(replacer_func, path)
        resolved_paths.append(resolved_path)
        if ALL_VAR_PAT.findall(resolved_path):
            for mt in ALL_VAR_PAT.findall(resolved_path):
                key = mt[1] or mt[2] or mt[3]
                if key not in var_dict: unresolved_vars.add(key) 
    return resolved_paths, list(unresolved_vars)

def inout_parse_exceptions_json(ajs_record, rules, bank, var_dict):
    def resolve_path(path_template, context_vars):
        path = path_template
        if "$" not in path: return path
        replacer_func = _create_replacer(context_vars)
        for _ in range(5): 
            last_path = path
            path = ALL_VAR_PAT.sub(replacer_func, path)
            path = re.sub(r"\$\{PM\[(\d+)\]\}", lambda m: ajs_record['param'].split()[int(m.group(1))] if len(ajs_record['param'].split()) > int(m.group(1)) else m.group(0), path)
            path = re.sub(r"\$\{EN\[([^}]+)\]\}", lambda m: var_dict.get(m.group(1), m.group(0)), path)
            if path == last_path: return path 
        return path

    for rule in rules:
        if not fnmatch(bank, rule.get("bank", "*")): continue
        if not fnmatch(os.path.basename(ajs_record['resource']), rule.get("shell", "*")): continue
        if not fnmatch(ajs_record['unit'], rule.get("unit", "*")): continue
            
        context_vars = copy.deepcopy(var_dict) 
        try: context_vars.update(dict(item.split('=', 1) for item in ajs_record['env'].split(';') if '=' in item))
        except Exception: pass
        try: 
            for i, p in enumerate(ajs_record['param'].split()): context_vars[f"{i+1}"] = p 
        except Exception: pass

        inputs = [resolve_path(p, context_vars) for p in rule.get("inputs", [])]
        outputs = [resolve_path(p, context_vars) for p in rule.get("outputs", [])]
        source_tag = rule.get("source_tag", "例外JSON")
        
        _, u_in = inout_resolve_path_variables(inputs, context_vars)
        _, u_out = inout_resolve_path_variables(outputs, context_vars)
        if u_in or u_out: source_tag = f"解析失敗: 未解決変数 (JSON) {{{', '.join(set(u_in + u_out))}}}"
        return inputs, outputs, source_tag
    return None, None, None

def inout_write_excel(path, records, headers):
    if not OPENPYXL_AVAILABLE: raise ImportError("openpyxl missing")
    NG_FILL = PatternFill(start_color="FFFFD2D2", end_color="FFFFD2D2", fill_type="solid")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AJS入出力解析"
    ws.append(headers)
    for record in records:
        in_files = "\n".join(record.get('inputs', []))
        out_files = "\n".join(record.get('outputs', []))
        ws.append([record.get('unit_full', ''), record.get('unit', ''), record.get('resource', ''), in_files, out_files, record.get('source_tag', '不明')])
        if "解析失敗" in record.get('source_tag', '') or (not in_files and not out_files and "リソース指定なし" not in record.get('source_tag', '')):
            for cell in ws[ws.max_row]: cell.fill = NG_FILL
    for row in ws.rows:
        for cell in row: cell.alignment = Alignment(wrap_text=True, vertical='top')
    for col in ws.columns:
        max_len = 0
        for cell in col:
            try: max_len = max(max_len, max(len(l) for l in str(cell.value).split('\n')))
            except: pass
        ws.column_dimensions[col[0].column_letter].width = min((max_len + 2) * 1.2, 70)
    wb.save(path)

def inout_write_csv(path, records, headers):
    with open(path, "w", encoding="utf-8-sig", newline="") as cf:
        w = csv.DictWriter(cf, fieldnames=headers)
        w.writeheader()
        for record in records:
            w.writerow({headers[0]: record.get('unit_full', ''), headers[1]: record.get('unit', ''), headers[2]: record.get('resource', ''), headers[3]: " | ".join(record.get('inputs', [])), headers[4]: " | ".join(record.get('outputs', [])), headers[5]: record.get('source_tag', '不明')})

# -----------------------------------------------------------------------------
# ★ 解析コアロジック (キャッシュ対応 & 変数名リファクタリング)
# -----------------------------------------------------------------------------
def analyze_ajs_jobs(gui_vars, gui_funcs, out_dir=None, use_cache=True):
    """
    AJS定義取得～I/O解析までを行う再利用可能な関数
    """
    global _ANALYSIS_CACHE, _LAST_CACHE_KEY
    
    update_status = gui_funcs['update_status']
    get_ssh_client = gui_funcs['get_ssh_client']
    show_error = gui_funcs['show_error']

    # Tab 3 (inout) or Tab 5 (dep)
    var_ajs = gui_vars.get('v_inout_ajs') or gui_vars.get('v_dep_ajs')
    ajs_path = var_ajs.get().strip() if var_ajs else ""

    var_res = gui_vars.get('v_inout_res') or gui_vars.get('v_dep_res')
    res_root = var_res.get().strip() if var_res else ""

    var_bank = gui_vars.get('v_inout_bank') or gui_vars.get('v_dep_bank')
    bank = var_bank.get() if var_bank else ""
    
    ajs_print_path = gui_vars['v_ajs_print_path'].get()
    jp1_hostname = gui_vars['v_jp1_hostname'].get()
    jp1_username = gui_vars['v_jp1_username'].get()

    current_key = (ajs_path, res_root, bank, jp1_hostname, jp1_username)
    
    if use_cache and current_key in _ANALYSIS_CACHE:
        update_status("解析結果をキャッシュから復元中...", 10)
        _log(f"[Cache] Hit! Using cached analysis data.")
        time.sleep(0.5) 
        return _ANALYSIS_CACHE[current_key]

    if not all([ajs_path, res_root, bank]):
        raise ValueError("AJSパス、リソースパス、銀行名が不足しています。")
        
    if out_dir is None:
        # 通常呼び出し時のフォールバック（基本は呼び出し元が作る）
        ts = time.strftime("%Y%m%d%H%M%S")
        out_dir = pathlib.Path(sys.argv[0]).resolve().parent / "temp_analysis" / ts
        out_dir.mkdir(parents=True, exist_ok=True)

    log_data = {
        "config_initial_vars": {}, "comenv_path": "未検出", "comenv_case_patterns": {}, 
        "comenv_master_dictionary": {}, "shell_cache_build_log": {}, "ajs_mapping": [], 
        "json_exceptions_log": {}, "shell_execution_log": {}, "ini_regex_log": {}, "final_records": []
    }
    
    _log(f"[Analyze] Start Analysis. Bank={bank}, Path={ajs_path}")

    initial_vars = {}
    if bank == "その他":
        custom_vars = gui_vars.get('v_inout_custom_vars', gui_vars.get('v_dep_custom_vars', []))
        initial_vars = dict(custom_vars)
    elif CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                initial_vars = json.load(f).get("comenv_initial_vars_by_bank", {}).get(bank, {})
        except Exception: pass
    log_data["config_initial_vars"] = initial_vars

    update_status("comenv 解析中...", 15)
    c_files = glob.glob(os.path.join(res_root, "**", "comenv"), recursive=True)
    comenv_path = c_files[0] if c_files else None
    log_data["comenv_path"] = comenv_path
    _log(f"[Analyze] comenv found: {comenv_path}")
    
    comenv_parser = ComenvParser(comenv_path, initial_vars, log_data)
    comenv_parser.parse_all_patterns()
    
    update_status("AJS定義取得中...", 30)
    export_list = []
    export_list.append(f'export JP1_HOSTNAME={shlex.quote(jp1_hostname)}')
    export_list.append(f'export JP1_USERNAME={shlex.quote(jp1_username)}')
    env_str = ' && '.join(export_list)
    
    # ★修正: 中間ファイルは tmp へ
    tmp_dir = out_dir / "tmp"
    tmp_dir.mkdir(exist_ok=True)
    
    remote_tmp = f"/tmp/ajs_out_anl_{time.time()}.txt"
    local_tmp = tmp_dir / "ajs_out_raw.txt" # tmp配下へ
    
    cmd = f'{env_str} && {ajs_print_path} -F AJSROOT1 -f "%JN%t%jn%t%sc%t%TY%t%En%t%pm" -R {shlex.quote(ajs_path)} > {remote_tmp}'
    
    _log(f"[Analyze] Executing remote command: {cmd}")
    
    with get_ssh_client() as ssh:
        ch = ssh.get_transport().open_session()
        ch.exec_command(cmd.encode("cp932"))
        if ch.recv_exit_status() != 0:
            err = ch.makefile_stderr().read().decode('cp932','ignore')
            _log(f"[Error] AJS command failed: {err}")
            raise RuntimeError(f"AJSコマンドエラー: {err}")
        sftp = ssh.open_sftp()
        sftp.get(str(remote_tmp), str(local_tmp))
        sftp.close()
        ssh.exec_command(f"rm -f {remote_tmp}")
        
    ajs_mapping_list = inout_parse_ajsprint_output(local_tmp)
    _log(f"[Analyze] Retrieved {len(ajs_mapping_list)} units.")
    log_data["ajs_mapping"] = ajs_mapping_list

    ex_rules = []
    if IO_EXCEPTION_FILE.exists():
        try:
            with open(IO_EXCEPTION_FILE, "r", encoding="utf-8") as f: ex_rules = json.load(f).get("rules", [])
        except: pass

    update_status("シェル解析キャッシュ構築...", 60)
    shell_cache = {}
    unique_shells = set(r['resource'] for r in ajs_mapping_list if r['resource'] and not r['resource'].endswith('.ini'))
    for sname in unique_shells:
        sfiles = glob.glob(os.path.join(res_root, "**", os.path.basename(sname)), recursive=True)
        if sfiles:
            procedures = ShellParser(sfiles[0]).get_procedures()
            shell_cache[sname] = procedures
            log_data["shell_cache_build_log"][sname] = procedures

    update_status("I/O変数解決実行中...", 70)
    final_records = []
    for record in ajs_mapping_list:
        r_copy = record.copy()
        inputs, outputs, tag = [], [], None
        var_dict = comenv_parser.get_var_dict_for_env(r_copy['env'])
        
        if not r_copy['resource']: tag = "リソース指定なし"
        
        if tag is None and ex_rules:
            i_j, o_j, t_j = inout_parse_exceptions_json(r_copy, ex_rules, bank, var_dict)
            if t_j:
                inputs, outputs, tag = i_j, o_j, t_j
                log_data["json_exceptions_log"][r_copy['unit_full']] = {'in': inputs, 'out': outputs, 'tag': tag}
        
        if tag is None and r_copy['resource'] in shell_cache:
            executor = ShellExecutor(shell_cache[r_copy['resource']], var_dict, r_copy)
            inputs, outputs, unres = executor.execute()
            tag = f"解析失敗: 未解決 {unres}" if unres else ("シェル解析 (IO定義無)" if not inputs and not outputs else "シェル解析 (変数解決)")
            log_data["shell_execution_log"][r_copy['unit_full']] = {'in': inputs, 'out': outputs, 'unresolved': unres, 'tag': tag}

        if tag is None and r_copy['resource'] and r_copy['resource'].endswith('.ini'):
            files = glob.glob(os.path.join(res_root, "**", os.path.basename(r_copy['resource'])), recursive=True)
            if files:
                raw_in, raw_out = inout_parse_ini_resource(files[0])
                inputs, u_in = inout_resolve_path_variables(raw_in, var_dict)
                outputs, u_out = inout_resolve_path_variables(raw_out, var_dict)
                tag = f"解析失敗: 未解決 {{{', '.join(set(u_in + u_out))}}}" if u_in or u_out else ("正規表現 (変数解決)" if inputs or outputs else "正規表現 (IO定義無)")
                log_data["ini_regex_log"][r_copy['unit_full']] = {'in': inputs, 'out': outputs, 'tag': tag}
            else: tag = "不明 (リソース無)"

        if tag is None: tag = "不明 (非解析対象)"
        
        r_copy['inputs'], r_copy['outputs'], r_copy['source_tag'] = inputs, outputs, tag
        final_records.append(r_copy)

    log_data["final_records"] = final_records
    
    _ANALYSIS_CACHE[current_key] = (final_records, log_data)
    _LAST_CACHE_KEY = current_key
    
    return final_records, log_data

def inout_start_job(gui_vars, gui_funcs):
    update_status = gui_funcs['update_status']
    save_hist = gui_funcs['save_hist']
    show_info = gui_funcs['show_info']
    show_error = gui_funcs['show_error']
    text_box = gui_vars.get('inout_text_box') 
    
    if text_box: text_box.delete('1.0', 'end')
    
    with open(LOG_FILE_RUN, "w", encoding="utf-8") as f:
        f.write(f"=== Tab 3 (InOut) Execution Start: {datetime.datetime.now()} ===\n")

    try:
        ts = time.strftime("%Y%m%d%H%M%S")
        base_dir = pathlib.Path(sys.argv[0]).resolve().parent
        out_dir = base_dir / DIR_NAME_INOUT / ts 
        out_dir.mkdir(parents=True, exist_ok=True)
        _log(f"[Info] Output Dir: {out_dir}")

        final_records, log_data = analyze_ajs_jobs(gui_vars, gui_funcs, out_dir, use_cache=False)

        out_format = gui_vars['v_inout_format'].get()
        headers = ["ユニット完全名称","ユニット名称","リソース名称","入力ファイル","出力ファイル","備考 (取得方法)"]
        update_status(f"{out_format}出力中...", 90)
        
        # ★修正: ファイル名固定化
        if out_format == "Excel":
            out_path = out_dir / "unit_io_mapping.xlsx"
            inout_write_excel(out_path, final_records, headers)
        else: 
            out_path = out_dir / "unit_io_mapping.csv"
            inout_write_csv(out_path, final_records, headers)
        
        _log(f"[Info] Saved result to: {out_path}")

        if text_box:
            problems = [f"・{r['unit_full']} ({r['source_tag']})" for r in final_records if "解析失敗" in r['source_tag'] or (not r['inputs'] and not r['outputs'] and "リソース指定なし" not in r['source_tag'])]
            if problems: text_box.insert('end', f"--- 問題検出 ({len(problems)}件) ---\n" + "\n".join(problems))
            else: text_box.insert('end', "--- 問題なし ---")

        write_detail_log(log_data)
        update_status("完了", 100)
        save_hist()
        _log("[Success] Completed.")
        show_info(f"解析が完了しました。\n結果は以下のファイルに出力されました:\n{out_path}")

    except Exception as e:
        tb = traceback.format_exc()
        _log(f"[Exception] {str(e)}\n{tb}")
        show_error(str(e))
    finally:
        update_status("待機中", 0)
